from fastapi import FastAPI, UploadFile, File
import os
import shutil
from openpyxl import load_workbook

from sqlalchemy import create_engine, Column, Integer, String
from sqlalchemy.orm import declarative_base, sessionmaker

app = FastAPI()

# -----------------------------
# Carpeta para guardar archivos
# -----------------------------
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# -----------------------------
# Base de datos SQLite
# -----------------------------
DATABASE_URL = "sqlite:///records.db"

engine = create_engine(
    DATABASE_URL,
    connect_args={"check_same_thread": False}
)

SessionLocal = sessionmaker(
    autocommit=False,
    autoflush=False,
    bind=engine
)

Base = declarative_base()


# -----------------------------
# Tabla de registros
# -----------------------------
class Record(Base):
    __tablename__ = "records"

    id = Column(Integer, primary_key=True, index=True)
    eori = Column(String)
    legal_name = Column(String)
    import_volume = Column(String)


Base.metadata.create_all(bind=engine)


# -----------------------------
# Endpoint principal
# -----------------------------
@app.get("/")
def home():
    return {"message": "API funcionando correctamente" }


# -----------------------------
# Subir y procesar Excel
# -----------------------------
@app.post("/upload")
async def upload_excel(file: UploadFile = File(...)):

    file_path = os.path.join(UPLOAD_FOLDER, file.filename)

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    workbook = load_workbook(file_path)
    sheet = workbook.active

    errors = []

    db = SessionLocal()

    for row_num in range(2, sheet.max_row + 1):

        eori = sheet[f"A{row_num}"].value
        legal_name = sheet[f"B{row_num}"].value
        import_volume = sheet[f"M{row_num}"].value

        # Ignorar filas vacías
        if not eori and not legal_name and import_volume is None:
            continue

        # Validaciones
        if not eori:
            errors.append({
                "row": row_num,
                "field": "EORI Number",
                "message": "Campo requerido"
            })

        if not legal_name:
            errors.append({
                "row": row_num,
                "field": "Declarant Legal Name",
                "message": "Campo requerido"
            })

        if import_volume is not None:
            if isinstance(import_volume, (int, float)):
                if import_volume < 0:
                    errors.append({
                        "row": row_num,
                        "field": "Import Volume",
                        "message": "Debe ser positivo"
                    })

        # Guardar registros válidos
        if eori and legal_name:
            record = Record(
                eori=str(eori),
                legal_name=str(legal_name),
                import_volume=str(import_volume)
            )

            db.add(record)

    db.commit()
    db.close()

    return {
        "message": "Archivo procesado correctamente",
        "filename": file.filename,
        "rows": sheet.max_row,
        "columns": sheet.max_column,
        "errors": errors
    }


# -----------------------------
# Consultar registros
# -----------------------------
@app.get("/records")
def get_records(page: int = 1, page_size: int = 20):

    db = SessionLocal()

    offset = (page - 1) * page_size

    records = db.query(Record).offset(offset).limit(page_size).all()

    result = []

    for record in records:
        result.append({
            "id": record.id,
            "eori": record.eori,
            "legal_name": record.legal_name,
            "import_volume": record.import_volume
        })

    total = db.query(Record).count()

    db.close()

    return {
        "page": page,
        "page_size": page_size,
        "total": total,
        "items": result
    }